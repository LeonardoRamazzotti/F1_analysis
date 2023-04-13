
import fastf1
import pandas as pd
import openpyxl
import numpy as np
from openpyxl import styles
from openpyxl.styles import  Color, PatternFill
from openpyxl.styles import numbers


def display_data(session,worksheet):

    #list of the driver  in the session
    driver_list=list(pd.unique(session.laps['Driver']))

    #car time parameter  avaiable
    index_list = ['Sector1Time','Sector2Time','Sector3Time','LapTime','Compound']

    #tyre color scheme for better recognition in worksheet
    color_dict= {'SOFT':'eb0e2b', 'MEDIUM':'00FFFF00','HARD':'ffffff','INTERMEDIATE':'0000FF00','WET':'000000FF'}
    
    #data display (cell insert) algorithm
    i=1
    worksheet.column_dimensions['B'].width= 11
    for  driver in driver_list:
        
        cell_merged='A'+str(i)
        worksheet[cell_merged]=driver
        j=i
        for category in index_list:
            worksheet['B'+str(j)]=category

            data_load=session.laps.pick_driver(driver)
            data_list=data_load[category] 
            
            k=3
            for  data in data_list:
                print(data)
                cell_data= worksheet.cell(row=j, column=k)
                cell_data.value=data
                cell_xy=cell_data.coordinate
                print(cell_xy)
                if category != 'Compound':
                    cell_data.number_format = "mm:ss.000"
                else:

                    # i had to use try -> except cause in the first lap in some session the tyre data is not registred
                    try:
                        worksheet[cell_xy].fill=PatternFill(fill_type='solid',fgColor=Color(color_dict[data]))
                    except:
                        worksheet[cell_xy].fill=PatternFill(fill_type='solid',fgColor=Color('66686b'))

                k+=1
 
            j+=1


        i+=6


def weekend_all_laps(year,gran_prix):

    workbook=openpyxl.Workbook()

    session_list = ['FP1','FP2','FP3','Q','R']

    for element in session_list:

        
        try:

            g_prix = fastf1.get_session(year,gran_prix,element)
            g_prix.load()

            # When you open a Workbook openpyxl autmatically create a worksheet, so you just have to  rename it.
            if element == 'FP1':
                sheet=workbook.active
                sheet.title= element
            else:
                sheet=workbook.create_sheet(element)
            
            display_data(g_prix,sheet)

        except:
            print("Session not loaded yet!")
        
    workbook.save('excels/'+str(year)+'_'+str(gran_prix)+'.xlsx')

    return



year = 2023
gp = "Australia"

try:
	fastf1.Cache.enable_cache('G:/My Drive/F1_Analysis/cache')
except:
    print('Cache not enable!')

weekend_all_laps(year,gp)